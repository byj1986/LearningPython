# coding=utf-8
import json
import openpyxl

from stockData import getCYWDailyData
from datetime import date, datetime, timedelta

xlsFile = r'D:\Projects\Python\LearningPython\StockMarket\股票主力成本.xlsx'
stockJson = r'D:\Projects\Python\LearningPython\StockMarket\stocks.json'


def getCWY(data, dayType: str) -> float:
    return round(data[dayType], 2)


def getTradeDate(data) -> date:
    tradeDate = datetime.strptime(data['TDate'], '%Y-%m-%dT%H:%M:%S')
    return tradeDate.date()


# def getExcelDate(date: str) -> date:
#     return datetime.strftime(date, '%Y-%m-%d %H:%M:%S')


def getExcelSheet(workBook, sheetName):
    sheet = None
    if sheetName in workBook.sheetnames:
        sheet = workBook[sheetName]
    else:
        sheet = workBook.create_sheet(sheetName)
        applyDefaultTemplate(sheet)
    return sheet


def applyDefaultTemplate(sheet):
    """
    Apply default template to the new sheet
    """
    sheet.cell(2, 1).value = '当日成本'
    sheet.cell(5, 1).value = '20日成本'
    sheet.cell(8, 1).value = '60日成本'
    sheet.cell(11, 1).value = '机构参与度'


def getRoundedNumber(rawData: any, precision: int = 2) -> float:
    return round(float(rawData), precision)


def updateExcel(stocks: any):
    for stock in stocks:
        # print(stock)
        cywData = getCYWDailyData(stock['id'])
        # print(cywData)
        tradeDate = getTradeDate(cywData)
        # print(tradeDate.ctime())
        # print(getTradeDate(cywData))
        stockName = cywData['Name']

        sheet = getExcelSheet(wb, stockName)
        columns = sheet.max_column
        lastDate = None
        if sheet.cell(1, columns).value:
            lastDate = sheet.cell(1, columns).value.date()

        # get last cell of row 1
        # if last cell equals cywData's date, update the cell
        # else create a new column
        columnNumber = columns+1
        if lastDate and lastDate == tradeDate:
            columnNumber = columns
        # print('Column number is ' + str(columnNumber))

        sheet.cell(1, columnNumber).value = tradeDate
        sheet.cell(1, columnNumber).data_type = 'd'
        sheet.cell(1, columnNumber)._style = dateStyle

        sheet.cell(2, columnNumber).value = getRoundedNumber(
            cywData['ZLCB'])
        sheet.cell(5, columnNumber).value = getRoundedNumber(
            cywData['ZLCB20R'])
        sheet.cell(8, columnNumber).value = getRoundedNumber(
            cywData['ZLCB60R'])
        sheet.cell(11, columnNumber).value = getRoundedNumber(
            cywData['JGCYD'])
        # break
        # Step 2: put data and update formula in excel

    # Step 3: save excel
    wb.save(xlsFile)


if __name__ == "__main__":
    wb = openpyxl.load_workbook(xlsFile)
    dateStyle = wb.worksheets[0].cell(1, 2)._style

    # Step 1: Get Data for yesterday for one stock
    with open(stockJson, 'r', encoding='utf-8') as stockFile:
        stocks = json.load(stockFile)
        updateExcel(stocks)
